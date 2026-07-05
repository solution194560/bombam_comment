#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
리디북스 작가 작품/댓글 수집기
────────────────────────────────────────────────────────────
■ 사용법
    python3 ridi_collector.py            # 메뉴 실행
    HEADLESS=0 python3 ridi_collector.py # 실제 창(최초 Cloudflare 통과용)

■ 메뉴
    1) 전체 수집   : 작가 페이지 → 작품목록 → 각 작품 전체댓글 → 엑셀
    2) 댓글만 수집 : 기존 작품목록(JSON)으로 댓글만 재수집 → 엑셀
    3) 설정 보기
    0) 종료
────────────────────────────────────────────────────────────
"""

import os
import re
import sys
import json
import time

# ════════════════════════════════════════════════════════════
#  ⚙️  설정 변수 (여기만 바꾸면 다른 작가에도 사용 가능)
# ════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════
#  설정(settings.json) — 텔레그램 봇에서 변경 가능. 파일이 없으면 아래 기본값 사용.
# ════════════════════════════════════════════════════════════
# settings.json 은 결과 폴더(OUTPUT_DIR, NAS에선 마운트된 /app/out)에 두어 변경값이 보존되게 함
SETTINGS_FILE = os.path.join(os.environ.get("OUTPUT_DIR", "."), "settings.json")

DEFAULT_SETTINGS = {
    "author":            "봄밤",   # 작가
    "author_url":        "https://ridibooks.com/author/117346?_s=search&_q=봄밤",
    "daily_time":        "11:00",  # 매일 실행 시간 (HH:MM)
    "notify_days":       2,         # 오늘 기준 며칠치 댓글을 알림 대상으로 볼지
    "notify_when_empty": False,     # 새 댓글 없을 때도 보낼지
    "last_run_date":     "",        # 마지막 실행 날짜(봇이 자동 기록)
    "subscribers":       [],        # 매일 자동 알림 수신자 chat_id 목록(봇에게 말 건 사람 자동 등록)
}


def load_settings():
    s = dict(DEFAULT_SETTINGS)
    if os.path.exists(SETTINGS_FILE):
        try:
            s.update({k: v for k, v in json.load(open(SETTINGS_FILE, encoding="utf-8")).items()
                      if k in DEFAULT_SETTINGS})
        except Exception as e:
            print(f"  [설정] settings.json 읽기 실패: {e}")
    return s


def save_settings(s):
    json.dump(s, open(SETTINGS_FILE, "w", encoding="utf-8"), ensure_ascii=False, indent=2)


_S = load_settings()

# ── 작가 / 주소 (settings.json 값) ─────────────────────
AUTHOR     = _S["author"]
AUTHOR_URL = _S["author_url"]

# ── 계정정보 (성인 콘텐츠/로그인 필요 시) ──────────────
#   · 별도 파일 account.json 에서 읽습니다.  (형식은 account.sample.json 참고)
#   · 환경변수(RIDI_ID / RIDI_PW / RIDI_LOGIN)가 있으면 그 값이 우선합니다.
ACCOUNT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "account.json")


def _load_account():
    """account.json + 환경변수에서 계정정보를 읽어온다 (환경변수 우선)."""
    acc = {"id": "", "password": "", "login": False}
    if os.path.exists(ACCOUNT_FILE):
        try:
            data = json.load(open(ACCOUNT_FILE, encoding="utf-8"))
            acc["id"] = str(data.get("id", "") or "")
            acc["password"] = str(data.get("password", "") or "")
            acc["login"] = bool(data.get("login", False))
        except Exception as e:
            print(f"  [경고] {os.path.basename(ACCOUNT_FILE)} 읽기 실패: {e}")
    if os.environ.get("RIDI_ID"):
        acc["id"] = os.environ["RIDI_ID"]
    if os.environ.get("RIDI_PW"):
        acc["password"] = os.environ["RIDI_PW"]
    if os.environ.get("RIDI_LOGIN"):
        acc["login"] = os.environ["RIDI_LOGIN"] == "1"
    return acc


ACCOUNT = _load_account()

# ── 동작 옵션 ─────────────────────────────────────────
HEADLESS       = os.environ.get("HEADLESS", "1") != "0"   # 0이면 실제 창
# 브라우저 프로필(로그인 쿠키 저장). 컨테이너로 이 폴더를 복사해 그대로 쓰면 로그인 재사용 가능.
#   기본: ~/.browser_data_dir  / 환경변수 BROWSER_PROFILE 로 변경 가능
PROFILE_DIR    = os.environ.get("BROWSER_PROFILE",
                                os.path.join(os.path.expanduser("~"), ".browser_data_dir"))
MAX_MORE_CLICK = 80          # '더보기' 최대 클릭 횟수
PER_WORK_DELAY = 0.3         # 작품 간 대기(초)

# 출력 폴더 (환경변수 OUTPUT_DIR 로 지정 가능. Synology 볼륨 마운트용. 기본: 현재 폴더)
OUTPUT_DIR   = os.environ.get("OUTPUT_DIR", ".")
os.makedirs(OUTPUT_DIR, exist_ok=True)

def _out(name):
    return os.path.join(OUTPUT_DIR, name)

# 출력 파일명 (작가명 기반)
WORKS_JSON   = _out(f"{AUTHOR}_작품목록.json")
REVIEWS_JSON = _out(f"{AUTHOR}_전체댓글.json")
EXCEL_OUT    = _out(f"{AUTHOR}_작품_전체댓글.xlsx")
LATEST_JSON  = _out(f"{AUTHOR}_작품별_최신댓글.json")
LATEST_EXCEL = _out(f"{AUTHOR}_작품별_최신댓글.xlsx")
RECENT_JSON  = _out(f"{AUTHOR}_최근30일_댓글.json")
RECENT_EXCEL = _out(f"{AUTHOR}_최근30일_댓글.xlsx")
RECENT_DAYS  = 30   # 3) 최근 N일 댓글 수집 기준

# ── 매일 텔레그램 알림 설정 (settings.json 값, 텔레그램 봇에서 변경 가능) ──
DAILY_TIME        = _S["daily_time"]         # 매일 실행 시간 (HH:MM)
NOTIFY_DAYS       = int(_S["notify_days"])   # 오늘 기준 며칠치 댓글을 볼지
NOTIFY_WHEN_EMPTY = bool(_S["notify_when_empty"])  # 새 댓글 없을 때도 보낼지

# ════════════════════════════════════════════════════════════
#  내부 로직
# ════════════════════════════════════════════════════════════

from playwright.sync_api import sync_playwright

# 작가 페이지에서 작품(제목+URL) 추출
WORKS_JS = r"""
() => {
  const out = [];
  const seen = new Set();
  document.querySelectorAll('a[href*="/books/"]').forEach(a => {
    const href = a.getAttribute('href') || '';
    const m = href.match(/\/books\/(\d+)/);
    if (!m) return;
    const id = m[1];
    if (seen.has(id)) return;
    let title = (a.innerText || a.textContent || '').trim();
    if (!title || title.length < 2) return;
    seen.add(id);
    out.push({
      title,
      url: 'https://ridibooks.com/books/' + id,
      book_id: id
    });
  });
  return out;
}
"""

# 리뷰 섹션 '더보기' JS 클릭
CLICK_MORE_JS = r"""
() => {
  const sec = document.querySelector('#ISLANDS__Review') ||
              document.querySelector('#detail_review');
  if (!sec) return false;
  const btns = [...sec.querySelectorAll('button')]
    .filter(e => (e.textContent || '').trim() === '더보기');
  if (btns.length) { btns[btns.length - 1].click(); return true; }
  return false;
}
"""

# 리뷰 추출 (날짜+작성자*** 둘 다 가진 li)
EXTRACT_JS = r"""
() => {
  const sec = document.querySelector('#ISLANDS__Review') ||
              document.querySelector('#detail_review');
  if (!sec) return {publish: null, total: null, reviews: []};
  const secText = sec.innerText || '';
  const mt = secText.match(/전체\s*([0-9,]+)/);
  const total = mt ? parseInt(mt[1].replace(/,/g, '')) : null;

  // 구매자 평점: 리뷰 li 안의 별 아이콘(viewBox 0 0 48 48) 중 '채워진 별' 개수.
  //  채워진 별은 빨강 rgb(229,76,67), 빈 별은 회색 rgb(230,230,230) → 색으로 판별(클래스 해시 비의존).
  const starRating = (li) => {
    const svgs = [...li.querySelectorAll('svg')].filter(s => s.getAttribute('viewBox') === '0 0 48 48');
    if (!svgs.length) return null;
    let filled = 0;
    for (const s of svgs) {
      const m = (getComputedStyle(s).color || '').match(/(\d+)\D+(\d+)\D+(\d+)/);
      if (!m) continue;
      const r = +m[1], g = +m[2], b = +m[3];
      const isGray = Math.abs(r-g) < 20 && Math.abs(g-b) < 20 && Math.abs(r-b) < 20;
      if (!isGray) filled++;   // 회색(빈 별)이 아니면 채워진 별
    }
    return filled;
  };

  const reviews = [];
  const seen = new Set();
  for (const li of sec.querySelectorAll('li')) {
    const txt = li.innerText || '';
    const dm = txt.match(/(\d{4}\.\d{2}\.\d{2})/);
    const um = txt.match(/([0-9A-Za-z]{2,}\*\*\*)/);
    if (!dm || !um) continue;
    const uIdx = txt.indexOf(um[1]);
    let content = txt.slice(0, uIdx).trim() || '(스포일러/내용없음)';
    const key = um[1] + '|' + dm[1] + '|' + content.slice(0, 15);
    if (seen.has(key)) continue;
    seen.add(key);
    reviews.push({content, user: um[1], date: dm[1], rating: starRating(li)});
  }

  let publish = null;
  const b = document.body.innerText;
  const pm = b.match(/(\d{4}\.\d{2}\.\d{2})\s*출간/) || b.match(/출간[^\d]*(\d{4}\.\d{2}\.\d{2})/);
  if (pm) publish = pm[1];
  return {publish, total, reviews};
}
"""

# 책 상세페이지에서 진짜 제목(og:title) 추출 — 로그인 없이도 동작
TITLE_JS = r"""
() => {
  const og = document.querySelector('meta[property="og:title"]');
  let t = og ? (og.content || '') : '';
  if (!t) { const h = document.querySelector('h1'); if (h) t = (h.innerText || '').trim(); }
  return (t || '').replace(/\s*[-|]\s*(리디|RIDI).*$/, '').trim();
}
"""


def _good_title(t):
    """수집된 제목이 '진짜 제목'인지 (placeholder/링크텍스트가 아닌지) 판단."""
    if not t:
        return False
    t = t.strip()
    bad = ("상세페이지 바로가기", "바로가기", "구독불가", "19세")
    return len(t) >= 2 and not any(b in t for b in bad)


def _wait_cloudflare(page, limit=90):
    """Cloudflare 'Just a moment' 화면이 사라질 때까지 대기 (진행 표시)."""
    for i in range(limit):
        if "Just a moment" not in (page.title() or ""):
            return True
        if i == 0:
            print("  [대기] Cloudflare 통과 대기 중...", end="", flush=True)
        else:
            print(".", end="", flush=True)
        time.sleep(1)
    print()
    return False


def _is_logged_in(page):
    """현재 로그인 상태인지 추정 (로그아웃/로그인 링크 유무로 판단)."""
    try:
        return page.evaluate(
            "() => !document.querySelector('a[href*=\"/account/login\"]')"
            " && !!document.querySelector('a[href*=\"logout\"], a[href*=\"/account/myidpassword\"], [class*=\"MyMenu\"]')"
        )
    except Exception:
        return False


def _try_login(page):
    """로그인 처리.
       1) 저장된 세션이 있으면 그대로 사용
       2) 자동 입력 시도
       3) 실패 & 창이 떠 있으면 → 직접 로그인 후 Enter (쿠키는 프로필에 저장돼 다음엔 자동)
    """
    if not ACCOUNT["login"]:
        print("  [로그인] 사용 안 함 (account.json login=false)")
        return
    try:
        page.goto("https://ridibooks.com/account/login", wait_until="domcontentloaded", timeout=60000)
    except Exception as e:
        print(f"  [로그인] 페이지 이동 실패: {str(e)[:50]}")
    _wait_cloudflare(page)

    # 1) 이미 로그인된 세션?
    if "login" not in (page.url or "") or _is_logged_in(page):
        print("  [로그인] 기존 세션 사용 (이미 로그인됨)")
        return

    # 2) 자동 입력 시도 (실패해도 멈추지 않도록 짧은 타임아웃)
    if ACCOUNT["id"] and ACCOUNT["password"]:
        try:
            print(f"  [로그인] {ACCOUNT['id']} 자동 입력 시도...")
            page.fill('input[name="user_id"], input#login_id, input[type="text"]',
                      ACCOUNT["id"], timeout=5000)
            page.fill('input[name="password"], input#login_pw, input[type="password"]',
                      ACCOUNT["password"], timeout=5000)
            page.locator('button[type="submit"], button:has-text("로그인")').first.click(timeout=5000)
            page.wait_for_load_state("domcontentloaded")
            time.sleep(2)
            _wait_cloudflare(page)
            if "login" not in (page.url or "") or _is_logged_in(page):
                print("  [로그인] 자동 로그인 성공 ✅")
                return
            print("  [로그인] 자동 로그인 미완료")
        except Exception as e:
            print(f"  [로그인] 자동 입력 실패: {str(e)[:50]}")

    # 3) 수동 로그인 (창이 떠 있고 + 사람이 있는 대화형 터미널일 때만)
    #    ※ 컨테이너/무인 서버(NAS)에서는 stdin 이 tty 가 아니므로 input() 으로
    #      영원히 멈추지 않도록 isatty() 로 방어. (자동 로그인 실패 시 hang 방지)
    if not HEADLESS and sys.stdin.isatty():
        print("\n  ──────────────────────────────────────────────")
        print("  ⚠️  열린 브라우저 창에서 직접 로그인하세요.")
        print("      (성인 콘텐츠는 로그인 후 자동 통과됩니다)")
        print("      로그인을 마친 뒤 이 터미널에서 [Enter] 를 누르세요.")
        print("  ──────────────────────────────────────────────")
        try:
            input("  로그인 완료 후 Enter > ")
        except EOFError:
            print("  (입력 불가 환경 → 30초 대기)")
            time.sleep(30)
        print("  [로그인] 진행합니다. (세션은 저장되어 다음엔 자동)")
    else:
        # 무인 환경(컨테이너/NAS) 또는 headless → 수동 로그인 불가. 멈추지 않고 넘어감.
        print("  ⚠️  자동 로그인 실패 + 무인 환경이라 수동 로그인 불가 → 이번 실행은 건너뜁니다.")
        print("      (맥에서 HEADLESS=0 으로 한 번 직접 로그인하면 세션이 profile 폴더에 저장돼 다음부터 자동)")


def _open_book_and_scroll(page, url):
    page.goto(url, wait_until="domcontentloaded", timeout=60000)
    _wait_cloudflare(page)
    if "adult-bridge" in page.url:
        try:
            page.locator('button:has-text("입장")').first.click(timeout=5000)
            page.wait_for_load_state("domcontentloaded")
            time.sleep(2)
        except Exception:
            pass
    for _ in range(6):
        page.evaluate("window.scrollBy(0, document.body.scrollHeight/5)")
        time.sleep(0.6)
    time.sleep(1.2)


def scrape_reviews_for_book(page, title, url, max_clicks=MAX_MORE_CLICK):
    """한 작품의 댓글 수집. max_clicks=0이면 '더보기' 없이 첫 화면만(빠름)."""
    _open_book_and_scroll(page, url)
    clicks = 0
    while clicks < max_clicks:
        try:
            if page.evaluate(CLICK_MORE_JS):
                clicks += 1
                time.sleep(1.0)
            else:
                break
        except Exception:
            break
    data = page.evaluate(EXTRACT_JS)
    data["_clicks"] = clicks
    try:
        data["page_title"] = page.evaluate(TITLE_JS)   # 상세페이지 진짜 제목
    except Exception:
        data["page_title"] = ""
    return data


def _latest_review(reviews):
    """리뷰 목록에서 날짜가 가장 최신인 댓글 1개 반환 (없으면 None)."""
    if not reviews:
        return None
    # 날짜(YYYY.MM.DD) 기준 내림차순, 같은 날짜면 먼저 나온 것
    return sorted(reviews, key=lambda r: r.get("date", ""), reverse=True)[0]


def _load_or_collect_works(page, tag=""):
    """작품목록 로드. 파일이 없거나 '0개'면 작가 페이지에서 다시 수집한다.
       (예전에 Cloudflare로 빈 목록이 저장돼 0개로 굳는 문제 방지)"""
    works = []
    if os.path.exists(WORKS_JSON):
        try:
            works = json.load(open(WORKS_JSON, encoding="utf-8"))
        except Exception:
            works = []
    # 결과폴더에 없으면, 이미지에 동봉된 기본 목록(스크립트 폴더)도 시도
    if not works:
        seed = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"{AUTHOR}_작품목록.json")
        if seed != WORKS_JSON and os.path.exists(seed):
            try:
                works = json.load(open(seed, encoding="utf-8"))
            except Exception:
                works = []
    if works:
        print(f"[{tag}] 기존 작품목록 {len(works)}개 사용")
        return works
    # 그래도 없으면 작가 페이지에서 수집
    print(f"[{tag}] 작품목록 없음/0개 → 작가 페이지에서 새로 수집")
    works = collect_work_list(page)
    if works:
        json.dump(works, open(WORKS_JSON, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    else:
        print(f"[{tag}] ⚠️ 작가 페이지에서도 0개 (로그인/Cloudflare 확인 필요) — 빈 목록은 저장 안 함")
    return works


def collect_work_list(page):
    """작가 페이지(+페이지네이션)에서 작품목록 수집."""
    print(f"[작품목록] {AUTHOR} - {AUTHOR_URL}")
    page.goto(AUTHOR_URL, wait_until="domcontentloaded", timeout=60000)
    _wait_cloudflare(page)
    all_works, seen = [], set()
    page_num = 1
    while True:
        for _ in range(5):
            page.evaluate("window.scrollBy(0, document.body.scrollHeight/4)")
            time.sleep(0.5)
        works = page.evaluate(WORKS_JS)
        new = [w for w in works if w["book_id"] not in seen]
        for w in new:
            seen.add(w["book_id"])
        all_works.extend(new)
        print(f"  [페이지 {page_num}] +{len(new)}개 (누계 {len(all_works)})")
        # 다음 페이지
        try:
            nxt = page.locator('a:has-text("다음"), button:has-text("다음")').first
            if nxt.count() and nxt.is_enabled():
                nxt.click(timeout=3000)
                time.sleep(2)
                _wait_cloudflare(page)
                page_num += 1
                continue
        except Exception:
            pass
        break
    return all_works


# ════════════════════════════════════════════════════════════
#  엑셀 생성
# ════════════════════════════════════════════════════════════

def build_excel(review_data, out_file):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # ── 최신 댓글 순 정렬: 작품은 최신 댓글 날짜 내림차순, 각 작품 내 댓글도 날짜 내림차순 ──
    def _maxdate(w):
        ds = [r.get("date", "") for r in w.get("reviews", [])]
        return max(ds) if ds else ""
    review_data = sorted(review_data, key=_maxdate, reverse=True)
    for w in review_data:
        w["reviews"] = sorted(w.get("reviews", []), key=lambda r: r.get("date", ""), reverse=True)

    rows = []
    for wi, w in enumerate(review_data, 1):
        title, pub, url = w["title"], w.get("publish_date") or "", w.get("url", "")
        revs = w.get("reviews", [])
        if not revs:
            rows.append({'번호': wi, '작품명': title, '출간일': pub, '댓글순위': '',
                         '작성자ID': '', '댓글 날짜': '', '댓글 내용': '(댓글 없음)', 'URL': url})
        else:
            for ri, r in enumerate(revs, 1):
                rows.append({
                    '번호': wi if ri == 1 else '', '작품명': title if ri == 1 else '',
                    '출간일': pub if ri == 1 else '', '댓글순위': ri,
                    '작성자ID': r.get('user', '-'), '댓글 날짜': r.get('date', ''),
                    '댓글 내용': r.get('content', ''), 'URL': url if ri == 1 else ''})
    df = pd.DataFrame(rows)
    summary = pd.DataFrame([{
        '번호': i, '작품명': w['title'], '출간일': w.get('publish_date') or '',
        '댓글수': len(w.get('reviews', [])), 'URL': w.get('url', '')
    } for i, w in enumerate(review_data, 1)])

    with pd.ExcelWriter(out_file, engine='openpyxl') as xl:
        summary.to_excel(xl, index=False, sheet_name='작품요약')
        df.to_excel(xl, index=False, sheet_name='전체댓글')

    wb = load_workbook(out_file)
    hf = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    hfill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft = Alignment(horizontal='left', vertical='center', wrap_text=True)
    link_font = Font(name='Arial', size=10, color='1155CC', underline='single')  # 작품명 링크
    thin = Side(border_style='thin', color='CCCCCC')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    # url_col = 작품명(B열) 하이퍼링크에 쓸 URL이 들어있는 열 (작업 후 삭제)
    for name, widths, leftcols, url_col in [
        ('작품요약', {'A': 6, 'B': 38, 'C': 12, 'D': 8}, {2}, 5),
        ('전체댓글', {'A': 6, 'B': 35, 'C': 12, 'D': 8, 'E': 12, 'F': 14, 'G': 70}, {2, 7}, 8),
    ]:
        ws = wb[name]
        # 작품명(B) → 상세페이지 하이퍼링크 (URL 열 값 사용), 적용 후 URL 열 삭제
        for ri in range(2, ws.max_row + 1):
            url = ws.cell(ri, url_col).value
            title_cell = ws.cell(ri, 2)
            if url and title_cell.value:
                title_cell.hyperlink = url
                title_cell.font = link_font
        ws.delete_cols(url_col)

        for c, wd in widths.items():
            ws.column_dimensions[c].width = wd
        for ci in range(1, ws.max_column + 1):
            cell = ws.cell(1, ci)
            cell.font, cell.fill, cell.alignment, cell.border = hf, hfill, ctr, bd
        ws.row_dimensions[1].height = 20
        for ri in range(2, ws.max_row + 1):
            fill = alt if ri % 2 == 0 else None
            for ci in range(1, ws.max_column + 1):
                cell = ws.cell(ri, ci)
                # 작품명 링크 폰트는 유지
                if not (ci == 2 and cell.hyperlink):
                    cell.font = Font(name='Arial', size=10)
                cell.alignment = lft if ci in leftcols else ctr
                cell.border = bd
                if fill:
                    cell.fill = fill
        ws.freeze_panes = 'A2'
    wb.save(out_file)


# ════════════════════════════════════════════════════════════
#  메뉴 동작
# ════════════════════════════════════════════════════════════

def _new_browser(p):
    mode = "headless(창 없음)" if HEADLESS else "실제 창"
    print(f"  [브라우저] {mode} 모드로 실행")
    return p.chromium.launch_persistent_context(
        user_data_dir=PROFILE_DIR, headless=HEADLESS,
        viewport={"width": 1280, "height": 900},
        # 저사양/컨테이너 안정화 플래그 (--disable-dev-shm-usage: /dev/shm 작은 환경 필수)
        args=['--no-sandbox', '--disable-setuid-sandbox', '--disable-dev-shm-usage',
              '--disable-gpu', '--start-maximized'])


def run_full():
    """1) 전체 수집: 작품목록 → 전체댓글 → 엑셀"""
    with sync_playwright() as p:
        b = _new_browser(p)
        page = b.new_page()
        _try_login(page)
        works = collect_work_list(page)
        if not works:
            print("⚠️  작품을 찾지 못했습니다. AUTHOR_URL을 확인하세요.")
            b.close()
            return
        json.dump(works, open(WORKS_JSON, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
        print(f"  → {WORKS_JSON} 저장 ({len(works)}개)\n")
        results = _collect_reviews(page, works)
        b.close()
    _finish(results)


def run_comments_only():
    """2) 댓글만 수집: 기존 작품목록 JSON 사용"""
    if not os.path.exists(WORKS_JSON):
        print(f"⚠️  {WORKS_JSON} 이 없습니다. 먼저 '1) 전체 수집'을 실행하세요.")
        return
    works = json.load(open(WORKS_JSON, encoding="utf-8"))
    print(f"[댓글만 수집] 기존 작품목록 {len(works)}개 사용")
    with sync_playwright() as p:
        b = _new_browser(p)
        page = b.new_page()
        _try_login(page)
        results = _collect_reviews(page, works)
        b.close()
    _finish(results)


def _collect_reviews(page, works):
    results = []
    for i, w in enumerate(works, 1):
        title, url = w["title"], w["url"]
        try:
            d = scrape_reviews_for_book(page, title, url)
            if _good_title(d.get("page_title")):      # 상세페이지의 진짜 제목으로 보정
                title = d["page_title"]
            revs = d.get("reviews", [])
            print(f"[{i:2d}/{len(works)}] {title[:26]:26s} | 댓글 {len(revs):3d}개 | 더보기 {d.get('_clicks',0)}회")
            results.append({"title": title, "url": url,
                            "publish_date": d.get("publish"),
                            "total_reviews": d.get("total"),
                            "reviews": revs})
        except Exception as e:
            print(f"[{i:2d}/{len(works)}] {title[:26]:26s} | ❌ {str(e)[:40]}")
            results.append({"title": title, "url": url, "reviews": [], "error": str(e)})
        json.dump(results, open(REVIEWS_JSON, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
        time.sleep(PER_WORK_DELAY)
    return results


def _finish(results):
    tot = sum(len(r.get("reviews", [])) for r in results)
    print(f"\n✅ 수집 완료: 작품 {len(results)}개 / 댓글 {tot}개")
    print(f"   JSON : {REVIEWS_JSON}")
    build_excel(results, EXCEL_OUT)
    print(f"   엑셀 : {EXCEL_OUT}")


# ════════════════════════════════════════════════════════════
#  작품별 '최신 댓글 1개'만 수집
# ════════════════════════════════════════════════════════════

def run_latest_only():
    """작품마다 가장 최신 댓글 1개(날짜+내용)만 수집 → 엑셀.
       작품목록이 없으면 작가 페이지에서 먼저 받아온다."""
    with sync_playwright() as p:
        b = _new_browser(p)
        page = b.new_page()
        _try_login(page)
        works = _load_or_collect_works(page, tag="최신댓글")
        results = []
        for i, w in enumerate(works, 1):
            title, url = w["title"], w["url"]
            try:
                d = scrape_reviews_for_book(page, title, url, max_clicks=0)  # 더보기 없이 첫 화면만
                if _good_title(d.get("page_title")):      # 상세페이지의 진짜 제목으로 보정
                    title = d["page_title"]
                latest = _latest_review(d.get("reviews", []))
                if latest:
                    print(f"[{i:2d}/{len(works)}] {title[:24]:24s} | 최신 {latest['date']} | {latest['content'][:18]}")
                else:
                    print(f"[{i:2d}/{len(works)}] {title[:24]:24s} | (댓글 없음)")
                results.append({"title": title, "url": url,
                                "publish_date": d.get("publish"),
                                "latest_date": latest["date"] if latest else "",
                                "latest_content": latest["content"] if latest else "(댓글 없음)",
                                "latest_user": latest["user"] if latest else ""})
            except Exception as e:
                print(f"[{i:2d}/{len(works)}] {title[:24]:24s} | ❌ {str(e)[:40]}")
                results.append({"title": title, "url": url, "latest_date": "",
                                "latest_content": f"(오류: {e})", "latest_user": ""})
            json.dump(results, open(LATEST_JSON, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
            time.sleep(PER_WORK_DELAY)
        b.close()

    # 최신 댓글 순 정렬 (날짜 내림차순, 댓글 없는 작품은 뒤로)
    results.sort(key=lambda r: r.get("latest_date", ""), reverse=True)
    build_latest_excel(results, LATEST_EXCEL)
    have = sum(1 for r in results if r.get("latest_date"))
    print(f"\n✅ 최신댓글 수집 완료: 작품 {len(results)}개 (댓글 있는 작품 {have}개)")
    print(f"   JSON : {LATEST_JSON}")
    print(f"   엑셀 : {LATEST_EXCEL}")


def build_latest_excel(results, out_file):
    """작품별 최신 댓글 1개 → 한 줄씩 엑셀."""
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    urls = [r.get('url', '') for r in results]
    df = pd.DataFrame([{
        '번호': i,
        '작품명': r['title'],
        '출간일': r.get('publish_date') or '',
        '최신 댓글 날짜': r.get('latest_date', ''),
        '최신 댓글 내용': r.get('latest_content', ''),
    } for i, r in enumerate(results, 1)])

    with pd.ExcelWriter(out_file, engine='openpyxl') as xl:
        df.to_excel(xl, index=False, sheet_name='작품별_최신댓글')

    wb = load_workbook(out_file)
    hf = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    hfill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    link_font = Font(name='Arial', size=10, color='1155CC', underline='single')  # 작품명 링크
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Side(border_style='thin', color='CCCCCC')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    ws = wb['작품별_최신댓글']
    for c, wd in {'A': 6, 'B': 40, 'C': 12, 'D': 14, 'E': 90}.items():
        ws.column_dimensions[c].width = wd
    for ci in range(1, ws.max_column + 1):
        cell = ws.cell(1, ci)
        cell.font, cell.fill, cell.alignment, cell.border = hf, hfill, ctr, bd
    ws.row_dimensions[1].height = 20
    for ri in range(2, ws.max_row + 1):
        fill = alt if ri % 2 == 0 else None
        for ci in range(1, ws.max_column + 1):
            cell = ws.cell(ri, ci)
            cell.font = Font(name='Arial', size=10)
            cell.alignment = lft if ci in {2, 5} else ctr
            cell.border = bd
            if fill:
                cell.fill = fill
        # 작품명(B열) → 상세페이지 하이퍼링크
        url = urls[ri - 2]
        if url:
            title_cell = ws.cell(ri, 2)
            title_cell.hyperlink = url
            title_cell.font = link_font
    ws.freeze_panes = 'A2'
    wb.save(out_file)


# ════════════════════════════════════════════════════════════
#  최근 N일(기본 30일) 댓글만 수집
# ════════════════════════════════════════════════════════════

def gather_recent(days):
    """오늘 기준 N일 내 댓글 행 리스트를 수집해 최신순으로 반환 (엑셀 생성 안 함).
       run_recent_days / 4_매일알림.py 가 공통으로 사용."""
    from datetime import datetime, timedelta
    cutoff = (datetime.now() - timedelta(days=days)).strftime("%Y.%m.%d")
    print(f"[최근 {days}일 댓글] 기준일: {cutoff} 이후")

    with sync_playwright() as p:
        b = _new_browser(p)
        page = b.new_page()
        _try_login(page)
        works = _load_or_collect_works(page, tag="최근댓글")

        rows = []
        for i, w in enumerate(works, 1):
            title, url = w["title"], w["url"]
            try:
                # 최근 댓글이 누락되지 않도록 전체 댓글을 받아 날짜로 거른다
                d = scrape_reviews_for_book(page, title, url)
                if _good_title(d.get("page_title")):      # 상세페이지의 진짜 제목으로 보정
                    title = d["page_title"]
                recent = [r for r in d.get("reviews", []) if r.get("date", "") >= cutoff]
                for r in recent:
                    rows.append({"title": title, "url": url, "publish_date": d.get("publish"),
                                 "date": r.get("date", ""), "user": r.get("user", ""),
                                 "rating": r.get("rating"), "content": r.get("content", "")})
                print(f"[{i:2d}/{len(works)}] {title[:24]:24s} | 최근 {days}일 댓글 {len(recent):3d}개")
            except Exception as e:
                print(f"[{i:2d}/{len(works)}] {title[:24]:24s} | ❌ {str(e)[:40]}")
            time.sleep(PER_WORK_DELAY)
        b.close()

    rows.sort(key=lambda r: r.get("date", ""), reverse=True)   # 최신 댓글 순
    return rows, cutoff


def format_notify_message(rows, days):
    """텔레그램 알림용 텍스트 구성 (최신 댓글 순)."""
    from datetime import datetime
    today = datetime.now().strftime("%-m/%-d")
    if not rows:
        return f"📭 {AUTHOR} — 최근 {days}일 새 댓글 없음 ({today})"
    lines = [f"📢 {AUTHOR} 최근 {days}일 새 댓글 {len(rows)}건", ""]
    for r in rows:
        content = (r.get("content") or "").replace("\n", " ").strip()
        if len(content) > 40:
            content = content[:40] + "…"
        pub = r.get("publish_date") or ""
        pub_str = f" ({pub} 출간)" if pub else ""     # 제목 옆: 출간일
        rating = r.get("rating")
        star = f"  ⭐{rating}" if isinstance(rating, int) and rating > 0 else ""   # 구매자 평점
        d = r.get("date", "")                          # 댓글 날짜 2026.07.01 → 07/01
        md = f"{d[5:7]}/{d[8:10]}" if len(d) >= 10 else d
        lines.append(f"💬 {r['title']}{pub_str}{star}")
        lines.append(f"{md} / “{content}”")             # 아래줄: 댓글날짜 / 댓글
        lines.append("")   # 항목 사이 빈 줄
    return "\n".join(lines).rstrip()


def run_recent_days(days=RECENT_DAYS):
    """최근 N일 동안 달린 댓글만 모아 최신순으로 엑셀 생성."""
    rows, cutoff = gather_recent(days)   # 이미 최신순 정렬됨
    json.dump(rows, open(RECENT_JSON, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    build_recent_excel(rows, RECENT_EXCEL, days, cutoff)
    print(f"\n✅ 최근 {days}일 댓글 수집 완료: 총 {len(rows)}개")
    print(f"   JSON : {RECENT_JSON}")
    print(f"   엑셀 : {RECENT_EXCEL}")


def build_recent_excel(rows, out_file, days, cutoff):
    """최근 N일 댓글 → 댓글 한 개당 한 줄, 최신순. 작품명은 상세페이지 링크."""
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    urls = [r.get("url", "") for r in rows]
    df = pd.DataFrame([{
        '번호': i,
        '댓글 날짜': r.get('date', ''),
        '작품명': r['title'],
        '작성자ID': r.get('user', ''),
        '댓글 내용': r.get('content', ''),
    } for i, r in enumerate(rows, 1)]) if rows else pd.DataFrame(
        columns=['번호', '댓글 날짜', '작품명', '작성자ID', '댓글 내용'])

    with pd.ExcelWriter(out_file, engine='openpyxl') as xl:
        df.to_excel(xl, index=False, sheet_name=f'최근{days}일_댓글')

    wb = load_workbook(out_file)
    hf = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    hfill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    link_font = Font(name='Arial', size=10, color='1155CC', underline='single')
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Side(border_style='thin', color='CCCCCC')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    ws = wb[f'최근{days}일_댓글']
    for c, wd in {'A': 6, 'B': 14, 'C': 38, 'D': 14, 'E': 95}.items():
        ws.column_dimensions[c].width = wd
    for ci in range(1, ws.max_column + 1):
        cell = ws.cell(1, ci)
        cell.font, cell.fill, cell.alignment, cell.border = hf, hfill, ctr, bd
    ws.row_dimensions[1].height = 20
    for ri in range(2, ws.max_row + 1):
        fill = alt if ri % 2 == 0 else None
        for ci in range(1, ws.max_column + 1):
            cell = ws.cell(ri, ci)
            cell.font = Font(name='Arial', size=10)
            cell.alignment = lft if ci in {3, 5} else ctr
            cell.border = bd
            if fill:
                cell.fill = fill
        # 작품명(C열) → 상세페이지 하이퍼링크
        url = urls[ri - 2]
        if url:
            title_cell = ws.cell(ri, 3)
            title_cell.hyperlink = url
            title_cell.font = link_font
    ws.freeze_panes = 'A2'
    wb.save(out_file)


def show_config():
    print(f"""
┌─ 현재 설정 ──────────────────────────────────
│ 작가(AUTHOR)      : {AUTHOR}
│ 주소(AUTHOR_URL)  : {AUTHOR_URL}
│ 계정 ID           : {ACCOUNT['id'] or '(미설정)'}
│ 계정 PW           : {'****' if ACCOUNT['password'] else '(미설정)'}
│ 로그인 사용       : {'예' if ACCOUNT['login'] else '아니오'}
│ HEADLESS          : {HEADLESS}  (0이면 실제 창)
│ 출력 엑셀         : {EXCEL_OUT}
└──────────────────────────────────────────────
""")


def menu():
    actions = {"1": run_full, "2": run_latest_only, "3": run_recent_days, "4": show_config}
    while True:
        print(f"""
╔══════════════════════════════════════════════╗
║   리디북스 댓글 수집기   (샘플 작가: {AUTHOR})
╠══════════════════════════════════════════════╣
║   1) 전체 수집        (작품목록 + 전체댓글 + 엑셀)
║   2) 작품별 최신댓글  (작품마다 최신 1개)
║   3) 최근 {RECENT_DAYS}일 댓글    (기간 내 댓글 최신순)
║   4) 설정 보기
║   0) 종료
╚══════════════════════════════════════════════╝
   ※ 모든 결과물은 최신 댓글 순으로 정렬됩니다.""")
        choice = input("선택 > ").strip()
        if choice == "0":
            print("종료합니다.")
            return
        fn = actions.get(choice)
        if fn:
            fn()
        else:
            print("잘못된 선택입니다.")


if __name__ == "__main__":
    # 인자로 바로 실행도 가능: full / latest / recent / comments / config
    arg = sys.argv[1] if len(sys.argv) > 1 else ""
    if arg == "full":
        run_full()
    elif arg == "latest":
        run_latest_only()
    elif arg == "recent":
        run_recent_days()
    elif arg == "comments":
        run_comments_only()
    elif arg == "config":
        show_config()
    else:
        menu()
